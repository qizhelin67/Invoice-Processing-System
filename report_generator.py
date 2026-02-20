"""
Excel Report Generator
Generate professional reimbursement reports from processed invoices
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from dataclasses import dataclass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl not available. Install with: pip install openpyxl")


class ExcelReportGenerator:
    """Generate Excel reimbursement reports"""

    # Category icons and names
    CATEGORY_INFO = {
        'taxi': {'icon': '🚕', 'name': '打车票'},
        'train': {'icon': '🚄', 'name': '火车飞机票'},
        'hotel': {'icon': '🏨', 'name': '住宿费'},
        'dining': {'icon': '🍜', 'name': '餐费'},
        'other': {'icon': '📦', 'name': '其他'}
    }

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        # Define styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.title_font = Font(name='Arial', size=16, bold=True, color='2F5597')
        self.subtotal_font = Font(name='Arial', size=11, bold=True)
        self.subtotal_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        self.total_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.total_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

    def generate_report(self,
                       organized_data: Dict[str, Any],
                       output_path: str) -> str:
        """
        Generate complete Excel report

        Args:
            organized_data: Organized invoice data from InvoiceOrganizer
            output_path: Where to save the Excel file

        Returns:
            Path to generated file
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        summary_sheet = wb.create_sheet('报销汇总', 0)
        detail_sheet = wb.create_sheet('报销明细', 1)

        # Generate summary sheet
        self._generate_summary_sheet(summary_sheet, organized_data)

        # Generate detail sheet
        self._generate_detail_sheet(detail_sheet, organized_data)

        # Save workbook
        wb.save(output_path)

        return output_path

    def _generate_summary_sheet(self, sheet, data: Dict[str, Any]):
        """Generate summary sheet with category totals"""

        stats = data['statistics']

        # Title
        sheet.merge_cells('A1:E1')
        title_cell = sheet['A1']
        title_cell.value = f"报销统计表 - {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_align

        # Headers
        headers = ['类别', '图标', '数量', '金额', '备注']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # Data rows
        row_num = 4
        for category, cat_stats in stats['by_category'].items():
            sheet.cell(row=row_num, column=1, value=cat_stats['name'])
            sheet.cell(row=row_num, column=2, value=cat_stats['icon'])
            sheet.cell(row=row_num, column=3, value=cat_stats['count'])
            sheet.cell(row=row_num, column=3).alignment = self.center_align
            sheet.cell(row=row_num, column=4, value=cat_stats['total_amount'])
            sheet.cell(row=row_num, column=4).number_format = '¥#,##0.00'
            sheet.cell(row=row_num, column=4).alignment = self.right_align

            # Apply borders
            for col in range(1, 6):
                sheet.cell(row=row_num, column=col).border = self.border

            row_num += 1

        # Total row
        sheet.cell(row=row_num, column=1, value='合计')
        sheet.cell(row=row_num, column=1).font = self.total_font
        sheet.cell(row=row_num, column=1).fill = self.total_fill
        sheet.cell(row=row_num, column=1).alignment = self.center_align

        sheet.cell(row=row_num, column=2).fill = self.total_fill

        sheet.cell(row=row_num, column=3, value=stats['total_invoices'])
        sheet.cell(row=row_num, column=3).font = self.total_font
        sheet.cell(row=row_num, column=3).fill = self.total_fill
        sheet.cell(row=row_num, column=3).alignment = self.center_align

        sheet.cell(row=row_num, column=4, value=stats['grand_total'])
        sheet.cell(row=row_num, column=4).font = self.total_font
        sheet.cell(row=row_num, column=4).fill = self.total_fill
        sheet.cell(row=row_num, column=4).number_format = '¥#,##0.00'
        sheet.cell(row=row_num, column=4).alignment = self.right_align

        # Apply borders to total row
        for col in range(1, 6):
            cell = sheet.cell(row=row_num, column=col)
            cell.border = self.border

        # Auto-fit columns
        self._auto_fit_columns(sheet)

        # Add metadata
        sheet.cell(row=row_num + 3, column=1, value='生成时间:')
        sheet.cell(row=row_num + 3, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row=row_num + 4, column=1, value='发票总数:')
        sheet.cell(row=row_num + 4, column=2, value=stats['total_invoices'])
        sheet.cell(row=row_num + 5, column=1, value='配对成功:')
        sheet.cell(row=row_num + 5, column=2, value=stats['total_pairs'])

    def _generate_detail_sheet(self, sheet, data: Dict[str, Any]):
        """Generate detailed sheet with all invoices"""

        categorized = data['categorized']
        pairs = data['pairs']

        # Title
        sheet.merge_cells('A1:G1')
        title_cell = sheet['A1']
        title_cell.value = f"报销明细表 - {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_align

        # Headers
        headers = ['类别', '文件名', '日期', '商家', '金额', '发票号', '备注']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # Data rows - group by category
        row_num = 4

        for category in ['taxi', 'train', 'hotel', 'dining', 'other']:
            invoices = categorized.get(category, [])

            if not invoices:
                continue

            # Category subtotal header
            cat_info = self.CATEGORY_INFO.get(category, self.CATEGORY_INFO['other'])
            sheet.merge_cells(f'A{row_num}:G{row_num}')
            cat_header = sheet.cell(row=row_num, column=1)
            cat_header.value = f"{cat_info['icon']} {cat_info['name']}"
            cat_header.font = Font(name='Arial', size=11, bold=True)
            cat_header.fill = self.subtotal_fill
            cat_header.alignment = self.left_align
            row_num += 1

            # Invoice rows
            for invoice in invoices:
                sheet.cell(row=row_num, column=1, value=cat_info['name'])
                sheet.cell(row=row_num, column=2, value=invoice.file_name)
                sheet.cell(row=row_num, column=3, value=invoice.date or '')
                sheet.cell(row=row_num, column=4, value=invoice.merchant or '')
                sheet.cell(row=row_num, column=5, value=invoice.amount or 0)
                sheet.cell(row=row_num, column=5).number_format = '¥#,##0.00'
                sheet.cell(row=row_num, column=5).alignment = self.right_align
                sheet.cell(row=row_num, column=6, value=invoice.invoice_number or '')

                # Check if paired
                pair_info = self._get_pair_info(invoice, pairs.get(category, []))
                sheet.cell(row=row_num, column=7, value=pair_info)

                # Apply borders
                for col in range(1, 8):
                    sheet.cell(row=row_num, column=col).border = self.border

                row_num += 1

            # Category subtotal
            cat_total = sum(inv.amount or 0 for inv in invoices)
            sheet.merge_cells(f'A{row_num}:F{row_num}')
            subtotal_cell = sheet.cell(row=row_num, column=1)
            subtotal_cell.value = f"小计: ¥{cat_total:,.2f} ({len(invoices)} 项)"
            subtotal_cell.font = self.subtotal_font
            subtotal_cell.fill = self.subtotal_fill
            subtotal_cell.alignment = self.right_align

            for col in range(1, 7):
                sheet.cell(row=row_num, column=col).border = self.border

            row_num += 2  # Extra space between categories

        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _get_pair_info(self, invoice, pairs: list) -> str:
        """Get pair information for invoice"""
        for pair in pairs:
            if pair.receipt == invoice:
                return f"已配对 ← {pair.invoice.file_name}"
            elif pair.invoice == invoice:
                return f"已配对 → {pair.receipt.file_name}"

        return ""

    def _auto_fit_columns(self, sheet):
        """Auto-fit column widths"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Add some padding
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width


def main():
    """Test report generator"""
    print("Excel Report Generator")
    print("=" * 50)

    # Sample data
    sample_data = {
        'categorized': {
            'taxi': [],
            'hotel': [],
        },
        'pairs': {},
        'statistics': {
            'total_invoices': 10,
            'total_pairs': 2,
            'grand_total': 1500.00,
            'by_category': {
                'taxi': {
                    'name': '打车票',
                    'icon': '🚕',
                    'count': 5,
                    'total_amount': 350.00,
                    'pairs': 2,
                    'unpaired': 1
                },
                'hotel': {
                    'name': '住宿费',
                    'icon': '🏨',
                    'count': 5,
                    'total_amount': 1150.00,
                    'pairs': 0,
                    'unpaired': 5
                }
            }
        }
    }

    generator = ExcelReportGenerator()
    print("\nTo generate a report:")
    print("  generator = ExcelReportGenerator()")
    print("  generator.generate_report(organized_data, 'output.xlsx')")


if __name__ == "__main__":
    main()
